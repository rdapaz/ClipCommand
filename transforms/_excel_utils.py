#!/usr/bin/env python3
"""
_excel_utils.py — cross-platform Excel population utility for ClipCommand.

Drop this file into your transforms/ folder alongside the excel_from_* scripts.
Import it in any transform:
    from _excel_utils import update_sheet, ExcelUtilsError

Public API
----------
update_sheet(data, sheet_name=None, start_cell="A1", workbook_path=None)
    Write *data* (list of lists) into the active (or named) sheet starting
    at *start_cell*.

    data           : list  — list of rows, each row a list of cell values
    sheet_name     : str   — worksheet name to target (None = active sheet)
    start_cell     : str   — top-left cell reference, default "A1"
    workbook_path  : str   — path to .xlsx (required for openpyxl fallback only;
                             ignored on Windows COM and macOS AppleScript paths,
                             which both operate on the currently active workbook)

Platform dispatch
-----------------
  Windows  → win32com COM automation  (live active workbook, no save needed)
  macOS    → AppleScript via osascript (live active workbook, no save needed)
  Linux /
  fallback → openpyxl                 (file on disk; workbook_path required;
                                       Excel must NOT have the file open)
"""

import os
import subprocess
import sys
from pathlib import Path


class ExcelUtilsError(Exception):
    """Raised for all excel_utils failures."""


# ── Cell reference helpers ────────────────────────────────────────────────────

def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """
    Convert "A1" → (row=1, col=1), "B3" → (row=3, col=2).
    Returns 1-based (row, col) tuple.
    """
    import re
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", cell_ref.strip())
    if not m:
        raise ExcelUtilsError(f"Invalid cell reference: {cell_ref!r}")
    col_str, row_str = m.group(1).upper(), m.group(2)
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(row_str), col


def _col_letter(col: int) -> str:
    """Convert 1-based column index to Excel letter(s): 1→A, 27→AA."""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


# ── Public entry point ────────────────────────────────────────────────────────

def update_sheet(data: list, sheet_name: str = None,
                 start_cell: str = "A1", workbook_path: str = None) -> str:
    """
    Write data into the target Excel sheet starting at start_cell.
    Returns a short status string suitable for use as transform() return value.
    """
    if not data:
        raise ExcelUtilsError("No data rows provided.")

    if sys.platform == "win32":
        return _update_via_com(data, sheet_name, start_cell)
    elif sys.platform == "darwin":
        return _update_via_applescript(data, sheet_name, start_cell)
    else:
        if not workbook_path:
            raise ExcelUtilsError(
                "workbook_path is required on Linux / non-Windows platforms."
            )
        return _update_via_openpyxl(data, sheet_name, start_cell, workbook_path)


# ── Windows: win32com ─────────────────────────────────────────────────────────

def _get_com_excel():
    """Return a live Excel.Application COM object, clearing gen_py cache if needed."""
    import shutil, re
    from win32com import client
    try:
        return client.gencache.EnsureDispatch("Excel.Application")
    except AttributeError:
        for mod in list(sys.modules):
            if re.match(r"win32com\.gen_py\..+", mod):
                del sys.modules[mod]
        shutil.rmtree(
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "Temp", "gen_py"),
            ignore_errors=True,
        )
        return client.gencache.EnsureDispatch("Excel.Application")


def _update_via_com(data: list, sheet_name: str, start_cell: str) -> str:
    try:
        app = _get_com_excel()
        app.Visible       = True
        app.DisplayAlerts = False
        wb = app.ActiveWorkbook
        if wb is None:
            raise ExcelUtilsError("No Excel workbook is currently open.")

        if sheet_name:
            try:
                ws = wb.Sheets(sheet_name)
            except Exception:
                raise ExcelUtilsError(
                    f"Sheet '{sheet_name}' not found in active workbook."
                )
        else:
            ws = wb.ActiveSheet

        start_row, start_col = _parse_cell_ref(start_cell)

        for r_offset, row in enumerate(data):
            for c_offset, value in enumerate(row):
                ws.Cells(start_row + r_offset, start_col + c_offset).Value = value

        sheet_label = sheet_name or ws.Name
        wb_label    = wb.Name
        return (
            f"OK — {len(data)} row(s) written to '{sheet_label}'!{start_cell} "
            f"in '{wb_label}' (Windows COM)"
        )
    except ExcelUtilsError:
        raise
    except Exception as exc:
        raise ExcelUtilsError(f"COM error: {exc}") from exc


# ── macOS: AppleScript ────────────────────────────────────────────────────────

def _update_via_applescript(data: list, sheet_name: str, start_cell: str) -> str:
    import json, tempfile

    start_row, start_col = _parse_cell_ref(start_cell)

    # Serialise data to temp JSON — sidesteps AppleScript quoting issues
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as tf:
        json.dump(data, tf, ensure_ascii=False)
        tmp_path = tf.name

    sheet_clause = (
        f'set ws to sheet "{sheet_name}" of wb'
        if sheet_name else
        "set ws to active sheet of wb"
    )

    script = f"""
use AppleScript version "2.4"
use scripting additions

set jsonPath to "{tmp_path}"
set pyCmd to "import sys,json; rows=json.load(open(sys.argv[1])); " & ¬
    "print('\\n'.join(['\\t'.join(str(c) for c in r) for r in rows]))"
set tabData to do shell script "python3 -c " & quoted form of pyCmd & " " & quoted form of jsonPath
set rowList to paragraphs of tabData

set startRow to {start_row}
set startCol to {start_col}

tell application "Microsoft Excel"
    if not (exists active workbook) then
        error "No Excel workbook is currently open."
    end if

    set wb to active workbook
    {sheet_clause}

    set rowIdx to startRow
    repeat with aRow in rowList
        set colIdx to startCol
        set cellValues to every text item of aRow
        repeat with aVal in cellValues
            set value of cell (rowIdx) of column colIdx of ws to (aVal as text)
            set colIdx to colIdx + 1
        end repeat
        set rowIdx to rowIdx + 1
    end repeat
end tell

return "OK"
"""

    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            raise ExcelUtilsError(f"AppleScript error: {result.stderr.strip()}")

        sheet_label = sheet_name or "active sheet"
        return (
            f"OK — {len(data)} row(s) written to '{sheet_label}'!{start_cell} "
            f"in active workbook (macOS AppleScript)"
        )
    except ExcelUtilsError:
        raise
    except subprocess.TimeoutExpired:
        raise ExcelUtilsError("AppleScript timed out after 60s.")
    except Exception as exc:
        raise ExcelUtilsError(f"osascript error: {exc}") from exc
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# ── Linux / fallback: openpyxl ────────────────────────────────────────────────

def _update_via_openpyxl(data: list, sheet_name: str,
                          start_cell: str, workbook_path: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ExcelUtilsError(
            "openpyxl is not installed. pip install openpyxl"
        )

    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise ExcelUtilsError(f"Workbook not found: {path}")

    wb = load_workbook(str(path))

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ExcelUtilsError(
                f"Sheet '{sheet_name}' not found in '{path.name}'. "
                f"Available: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    start_row, start_col = _parse_cell_ref(start_cell)

    for r_offset, row in enumerate(data):
        for c_offset, value in enumerate(row):
            ws.cell(
                row=start_row + r_offset,
                column=start_col + c_offset,
                value=value
            )

    wb.save(str(path))
    sheet_label = sheet_name or ws.title
    return (
        f"OK — {len(data)} row(s) written to '{sheet_label}'!{start_cell} "
        f"in '{path.name}' (openpyxl, saved to disk)"
    )
